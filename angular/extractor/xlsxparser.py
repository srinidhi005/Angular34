import glob
import os
import sys

import openpyxl
import re
import config
import pdftables_api

API_KEY = '6yc5mavlgq4v'
with open("hello.txt","a") as op:
    #op.write("executed\n")
units_pattern = config.units_pattern
company_names_pattern = config.company_names_pattern
statements = config.statements
sga_patterns = config.sga_patterns
cost_of_goods_sold=config.cost_of_goods_sold
total_revenue = config.total_revenue
research_and_dev = config.research_and_dev
product_dev = config.product_dev
sales_and_marketing = config.sales_and_marketing
total_operating_exp = config.total_operating_exp
net_interest_exp = config.net_interest_exp
other_income_exp = config.other_income_exp
tax_patterns = config.tax_patterns

def get_all_excel_files(pdf_folder_path):
    xlsx_files = glob.glob(pdf_folder_path + "/*.xlsx")  # collect all xlsx files
    return xlsx_files

def get_company_name_and_units_used(xlsx_files):
    """
    :param pdf_folder_path:
    :return: dictionary containing company and units used
    """
    company_names = []
    units_used = []

    for xl_path in xlsx_files: # loop through all xlsx files
        path, filename = os.path.split(xl_path) # split file and path
        work_book = openpyxl.load_workbook(xl_path) # load workbook
        no_of_sheets = len(work_book.sheetnames) #get sheets count
        for sheet_no in range(no_of_sheets): # loop through all sheets
            work_sheet = work_book.worksheets[sheet_no] # take sheet reference
            found = 0
            for j in range(1,4): #search upto 3rd row
                for i in range(1,4): #search upto 3rd column to get company name
                    if work_sheet.cell(row=j,column=i).value is not None: # check non empty cells only
                        for com_name_pattern in company_names_pattern: # loop through all possible copany names pattern
                            match_obj = re.match(com_name_pattern,str(work_sheet.cell(row=j,column=i).value),re.I) #match with ignoring case
                            if match_obj:
                                company_names.append(match_obj.group(0))
                                found = 1
                                break
                    if found == 1:
                        break
            if found == 1:
                break


    # for company in company_names:
    #     print(company)
    file_names = []
    for xl_path in xlsx_files: # loop through all xlsx files
        path, filename = os.path.split(xl_path) # split file and path
        work_book = openpyxl.load_workbook(xl_path) # load workbook
        no_of_sheets = len(work_book.sheetnames) #get sheets count
        for sheet_no in range(no_of_sheets): # loop through all sheets
            work_sheet = work_book.worksheets[sheet_no] # take sheet reference
            found = 0
            for j in range(1,7): #search upto 6th row
                for i in range(1,4): #search upto 3rd column to get units used
                    if work_sheet.cell(row=j,column=i).value is not None:
                        for unit_pattern in units_pattern:
                            match_obj = re.search(unit_pattern,str(work_sheet.cell(row=j,column=i).value),re.I)
                            if match_obj:
                                units_used.append(match_obj.group(0))
                                found = 1
                                break
                    if found == 1:
                        break
            if found == 1:
                break
        file_names.append(filename)
    return dict(zip(company_names,units_used)),dict(zip(file_names,company_names))

def get_required_sheets(xlsx_files):
    """

    :return: returns sheet references
    """
    sheet_ref = {}
    for xl_path in xlsx_files: # loop through all xlsx files
        path, filename = os.path.split(xl_path) # split file and path
        # print(filename)
        work_book = openpyxl.load_workbook(xl_path) # load workbook
        no_of_sheets = len(work_book.sheetnames) #get sheets count
        for sheet_no in range(no_of_sheets): # loop through all sheets
            work_sheet = work_book.worksheets[sheet_no] # take sheet reference
            # print(work_sheet.title,"in this page")
            found = 0
            for j in range(1,8): #search upto 3rd row
                for i in range(1,5): #search upto 3rd column to get statements name
                    if work_sheet.cell(row=j, column=i).value is not None:
                        for statement in statements:
                            # print(statement)
                            match_obj = re.match(statement, str(work_sheet.cell(row=j, column=i).value), re.I|re.MULTILINE)
                            if match_obj:
                                # print(work_sheet.title)
                                sheet_ref[xl_path] = work_sheet.title
                                found = 1
                                break
                if found == 1:
                    break
            if found == 1:
                break

    return sheet_ref

def check_multi_cell_and_split(multiline_str):
    # print("multiline cell function called")
    terms = re.findall("[A-Z][^A-Z]*",multiline_str)
    return terms

def get_years(xlsx_files,filename_company):
    """
    fetch years and retueen comapny  year dict
    :return:
    """
    years_dict = {}
    for xl_path, page_ref in xlsx_files.items():  # loop through all xlsx files
        path, filename = os.path.split(xl_path)  # split file and path
        work_book = openpyxl.load_workbook(xl_path)  # load workbook
        work_sheet = work_book[page_ref]  # take sheet reference

        flg = 1
        for j in range(1, 10):
            count = 0
            years = []
            for i in range(1, 10):
                if work_sheet.cell(row=j, column=i).value is not None:
                    match_obj = re.match("^2\d{3}$", str(work_sheet.cell(row=j, column=i).value), re.I)
                    if match_obj:
                        count += 1
                        # print(match_obj.group(0),"year")
                        years.append(match_obj.group(0))
                    if count == 3:
                        flg = 0
                        break
            years_dict[filename_company[filename]] = years
            if flg == 0:
                break

    return years_dict

def get_required_values(xlsx_files,search_term,filename_company):
    """
    :return: returns sga and total revenue

    apple is having empty. its avilable in next
    """

    term_patterns = search_term
    term_dict = {}
    for xl_path,page_ref in xlsx_files.items(): # loop through all xlsx files
        path, filename = os.path.split(xl_path) # split file and path
        # print(filename)
        work_book = openpyxl.load_workbook(xl_path) # load workbook
        work_sheet = work_book[page_ref]  # take sheet reference
        found = 0
        row_count = work_sheet.max_row
        column_count = work_sheet.max_column

        for j in range(1, row_count):  # search upto last row
            for i in range(1, column_count):  # search last coulmn
                if work_sheet.cell(row=j, column=i).value is not None:
                    for term in term_patterns:
                        match_obj = re.search(term, str(work_sheet.cell(row=j, column=i).value), re.I)
                        if match_obj:
                            # print(str(work_sheet.cell(row=j, column=i).value))
                            term_values = []
                            terms_found = check_multi_cell_and_split(work_sheet.cell(row=j, column=i).value)
                            if len(terms_found) == 1:
                                if re.match(term,terms_found[0],re.I):
                                    for k in range(10):
                                        sga_match_obj = re.search("\d+[\.\d+].*",str(work_sheet.cell(row=j, column=k+2).value))
                                        if sga_match_obj:
                                            # print(str(work_sheet.cell(row=j, column=k+2).value))
                                            term_values.append(str(work_sheet.cell(row=j, column=k+2).value))
                                            found = 1
                                    term_dict[filename_company[filename]] = term_values
                                    if found == 1:
                                        break

                            elif len(terms_found) > 1:
                                print("Multiline found",terms_found)
                                for k in range(10):
                                    no_of_values = re.findall("[1-9,.]+",
                                                              str(work_sheet.cell(row=j, column=k + 2).value))
                                    if no_of_values != []:
                                        # print(str(work_sheet.cell(row=j, column=k+2).value))

                                        for key,value in zip(terms_found,no_of_values):
                                            if re.match(term,key,re.I):
                                                term_values.append(value)
                                                break

                                        found = 1
                                term_dict[filename_company[filename]] = term_values
                                if found == 1:
                                    break


                    if found == 1:
                        break

    return term_dict

def convert_to_thousand(terms,unit_dict):
    """
    if values are in terms of Million then Multiply by 1000
    to make it in thousands
    :return:
    """
    temp_dict = {}
    for key,value in terms.items():
        modified_value = []
        for val in value:
            if re.search("\(",val):
                val = val.strip("(") # repmove (
                val = val.strip(")") # remove )
                val = val.replace(",","") # remove comma
                val = -float(val) # conver to float and minus it
            else:
                val = val.replace(",", "") # remove comma
                val = float(val) #conver to float and minus it
            modified_value.append(val)
        if key in unit_dict.keys():
            if unit_dict[key].lower() == "millions":
                modified_value = [val * 1000 for val in modified_value]
            else:
                modified_value = modified_value

        temp_dict[key] = modified_value
    return temp_dict

def convert_pdf_to_excel(pdf_path):
    c = pdftables_api.Client(API_KEY)

    path, file = os.path.split(pdf_path)
    excelfile = file.split(".")[-2]
    excelfile_path = "./temp"
    c.xlsx(pdf_path,excelfile_path+"/temporary")

    return excelfile_path

if __name__ == "__main__":

    pdf_path = sys.argv[1]

    pdf_file_path = r"C:\pdf\output\Nike.pdf"  # excel files path
    pdf_file_path = pdf_path  # excel files path
    excel_files_path = convert_pdf_to_excel(pdf_file_path)
    excel_files = get_all_excel_files(excel_files_path)
    company_and_unit,filename_company = get_company_name_and_units_used(excel_files)
    print("COMPANY Units",company_and_unit)
    sheet_ref = get_required_sheets(excel_files)

    company_years = get_years(sheet_ref,filename_company)
    print("Company Years",len(company_years),company_years)

    sga = get_required_values(sheet_ref,sga_patterns,filename_company)
    modified_sga = convert_to_thousand(sga,company_and_unit)
    print("SGA",len(modified_sga),modified_sga)

    total_revenue = get_required_values(sheet_ref,total_revenue,filename_company)
    modified_total_revenue = convert_to_thousand(total_revenue, company_and_unit)
    print("Total Revenue", len(modified_total_revenue),modified_total_revenue)

    research_and_dev = get_required_values(sheet_ref, research_and_dev,filename_company)
    mod_research_and_dev = convert_to_thousand(research_and_dev,company_and_unit)
    print("Research & Development",len(mod_research_and_dev),mod_research_and_dev)

    product_dev = get_required_values(sheet_ref, product_dev,filename_company)
    mod_prod_dev = convert_to_thousand(product_dev,company_and_unit)
    print("Product & Development", len(mod_prod_dev.keys()), mod_prod_dev)

    sales_and_marketing = get_required_values(sheet_ref, sales_and_marketing,filename_company)
    mod_sales_and_marketing = convert_to_thousand(sales_and_marketing,company_and_unit)
    print("Sales & Marketing", len(mod_sales_and_marketing.keys()), mod_sales_and_marketing)


    total_operating_exp = get_required_values(sheet_ref, total_operating_exp,filename_company)
    mod_total_operating_exp = convert_to_thousand(total_operating_exp,company_and_unit)
    print("Total Operating Expense", len(mod_total_operating_exp.keys()), mod_total_operating_exp)

    net_interest_exp = get_required_values(sheet_ref, net_interest_exp,filename_company)
    net_interest_exp = convert_to_thousand(net_interest_exp,company_and_unit)
    print("Net Interest Expense", len(net_interest_exp.keys()), net_interest_exp)
    #
    other_income_exp = get_required_values(sheet_ref, other_income_exp,filename_company)
    other_income_exp = convert_to_thousand(other_income_exp,company_and_unit)
    print("Other Income Expense", len(other_income_exp.keys()), other_income_exp)

    tax_val_dict = get_required_values(sheet_ref,tax_patterns,filename_company)
    tax_val_dict = convert_to_thousand(tax_val_dict,company_and_unit)
    print("Taxes",len(tax_val_dict.keys()),tax_val_dict)

    cogs = get_required_values(sheet_ref,cost_of_goods_sold,filename_company)
    cogs = convert_to_thousand(cogs,company_and_unit)
    print("COGS",len(cogs.keys()),cogs)
    op.write("Program Executed Succesfully\n")




